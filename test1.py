from openpyxl import *
import xlrd
import pymongo
import os
import string

myclient = pymongo.MongoClient('mongodb://localhost:27017/')
 
dblist = myclient.list_database_names()
if "cet" in dblist:
  print("Database already exists!")

mydb = myclient["cet"]
mycol = mydb["cet4"]


def read_excel_to_mongodb():
    workbook = xlrd.open_workbook(r'cet4.xls')
    print(workbook.sheet_names())
    sheet = workbook.sheet_by_name('Sheet1')
    ncols = sheet.ncols
    nrows = sheet.nrows
    print(nrows, ncols)
    # return 0
    for n in range(1, nrows):
        word_dict = { "word": sheet.cell(n, 1).value[:-1], 
        "phonetic_notation": sheet.cell(n, 2).value, 
        "paraphrase": sheet.cell(n, 3).value[1:] }
        x = mycol.insert_one(word_dict) 
        print(n, word_dict)
    x = mycol.find_one()
    print(x)

        

def main():
    read_excel_to_mongodb()


if __name__ == "__main__":
    main()
